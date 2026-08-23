"""Saved responses routes.

Security:
- Feature-gated to Plus/Pro.
- Field lengths aligned with DB columns so inserts cannot 500 on oversize.
- Per-user cap + rate limit so authenticated spam cannot fill saved_responses.
- Delete uses scoped lookup (404 for missing *and* foreign rows) so IDs
  cannot be enumerated via 403 vs 404.

Functionality:
- GET /saved supports search (prompt + one_liner substring), persona_id
  filter, score filter (min_score), pagination, and sort modes (newest /
  oldest / score / pinned).
- DELETE /saved/bulk accepts a JSON list of ids for one-shot cleanup.
"""

import csv
import io
import json
import logging
from typing import Optional

from pydantic import BaseModel, Field, field_validator
from fastapi import APIRouter, Depends, HTTPException, Query
from sqlalchemy import or_
from sqlalchemy.orm import Session

try:
    import openpyxl
    from openpyxl import Workbook
    OPENPYXL_AVAILABLE = True
except ImportError:
    OPENPYXL_AVAILABLE = False

from arena.core.dependencies import get_current_user_required
from arena.core.input_validation import sanitize_model_optional_text, sanitize_model_text
from arena.core.rate_limits import enforce_user_rate_limit
from arena.core.tier_config import get_tier_str, has_feature, normalize_tier
from arena.core.datetime_utils import utcnow_naive
from arena.database import get_db
from arena.db_models import SavedResponse
from arena.models.schemas import UserResponse

logger = logging.getLogger(__name__)

router = APIRouter(tags=["saved"])

# Hard cap on stored takes per user (UI already treats this as a personal list).
SAVED_MAX_PER_USER = 200

# Bulk delete cap — even a power user shouldn't be able to wipe their whole
# library in a single click; 50 is enough for the "select all visible" UI
# pattern without becoming a footgun.
SAVED_BULK_DELETE_MAX = 50

# Pin cap — pinned takes are meant to be a small curated set at the top of
# the sidebar, not a second list. 50 keeps the query cheap and the UI honest.
SAVED_PIN_MAX = 50


def _escape_like(value: str) -> str:
    """Escape SQL LIKE wildcards. % and _ are wildcards; without escaping,
    a user typing '100%' would match every row."""
    return (
        value.replace("\\", "\\\\")
        .replace("%", "\\%")
        .replace("_", "\\_")
    )


def normalize_export_search(value):
    """Tolerantly normalize a search term for export/preview queries.

    Write-time endpoints sanitize strictly (422 on junk), but presets saved
    before that rule can hold whitespace-only or otherwise invalid search
    values. Those must degrade to "no search filter" in the shared query
    builder instead of raising a 500 for legacy rows.
    """
    if value is None:
        return None
    try:
        return sanitize_model_optional_text(
            value, max_length=100, field_name="search"
        )
    except ValueError:
        return None


def build_saved_export_query(
    db: Session,
    user_id: int,
    *,
    search: Optional[str],
    persona_id: Optional[str],
    min_score: Optional[int],
    max_score: Optional[int],
    sort: str,
    pinned: Optional[bool] = None,
):
    """Build the saved-response query shared by exports and previews.

    Single source of truth for export filters/sort so any "dry run"
    (e.g. export preset preview) counts and samples exactly what a real
    export returns. Callers must scope results to the owning user id.
    """
    q = db.query(SavedResponse).filter(SavedResponse.user_id == user_id)

    safe_search = normalize_export_search(search)
    if safe_search:
        escaped = _escape_like(safe_search)
        pattern = f"%{escaped}%"
        q = q.filter(
            or_(
                SavedResponse.prompt.ilike(pattern, escape="\\"),
                SavedResponse.one_liner.ilike(pattern, escape="\\"),
            )
        )

    if persona_id:
        q = q.filter(SavedResponse.persona_id == persona_id)

    if min_score is not None:
        q = q.filter(SavedResponse.score >= min_score)

    if max_score is not None:
        q = q.filter(SavedResponse.score <= max_score)

    if pinned is not None:
        if pinned:
            q = q.filter(SavedResponse.pinned_at.isnot(None))
        else:
            q = q.filter(SavedResponse.pinned_at.is_(None))

    if sort == "oldest":
        q = q.order_by(SavedResponse.saved_at.asc(), SavedResponse.id.asc())
    elif sort == "pinned":
        q = q.order_by(
            SavedResponse.pinned_at.desc().nullslast(),
            SavedResponse.saved_at.desc(),
            SavedResponse.id.desc(),
        )
    elif sort == "score":
        q = q.order_by(SavedResponse.score.desc().nullslast(), SavedResponse.id.desc())
    else:  # newest (default)
        q = q.order_by(SavedResponse.saved_at.desc(), SavedResponse.id.desc())

    return q


class SavedRequest(BaseModel):
    session_id: str = Field(..., min_length=1, max_length=36)
    agent_id: str = Field(..., min_length=1, max_length=20)
    persona_id: str = Field(..., min_length=1, max_length=50)
    persona_name: str = Field(..., min_length=1, max_length=255)
    persona_color: str = Field(..., min_length=1, max_length=20)
    prompt: str = Field(..., min_length=1, max_length=1000)
    one_liner: str = Field(..., min_length=1, max_length=1000)
    # Text column — still bound so a single body cannot dump multi-MB prose
    # if the global request-size middleware is ever raised for this path.
    verdict: str = Field(..., min_length=1, max_length=20000)
    score: int | None = Field(None, ge=0, le=100)
    confidence: int | None = Field(None, ge=0, le=100)

    @field_validator(
        "session_id",
        "agent_id",
        "persona_id",
        "persona_name",
        "persona_color",
        "prompt",
        "one_liner",
        "verdict",
    )
    @classmethod
    def strip_required(cls, v: str, info) -> str:
        # persona_name / color may include spaces; use text sanitizer not html strip.
        max_len = {
            "session_id": 36,
            "agent_id": 20,
            "persona_id": 50,
            "persona_name": 255,
            "persona_color": 20,
            "prompt": 1000,
            "one_liner": 1000,
            "verdict": 20000,
        }[info.field_name]
        return sanitize_model_text(v, max_length=max_len, field_name=info.field_name)


class BulkDeleteRequest(BaseModel):
    """Body schema for DELETE /saved/bulk. IDs not owned by the caller
    are silently ignored (no existence oracle)."""
    ids: list[int] = Field(..., min_length=1, max_length=SAVED_BULK_DELETE_MAX)


class PinRequest(BaseModel):
    """Body schema for PATCH /saved/{id} — pin or unpin one saved take."""
    pinned: bool


class BulkPinRequest(BaseModel):
    """Body schema for PATCH /saved/bulk-pin — pin or unpin many takes."""
    ids: list[int] = Field(..., min_length=1, max_length=SAVED_MAX_PER_USER)
    pinned: bool


@router.get("/saved")
async def get_saved(
    user: UserResponse = Depends(get_current_user_required),
    db: Session = Depends(get_db),
    page: int = Query(1, ge=1),
    per_page: int = Query(50, ge=1, le=SAVED_MAX_PER_USER),
    search: Optional[str] = Query(None, max_length=100, description="Case-insensitive substring match on prompt + one_liner."),
    persona_id: Optional[str] = Query(None, max_length=50, description="Restrict to one persona."),
    min_score: Optional[int] = Query(None, ge=0, le=100, description="Minimum score (inclusive)."),
    max_score: Optional[int] = Query(None, ge=0, le=100, description="Maximum score (inclusive)."),
    pinned: Optional[bool] = Query(None, description="Restrict to pinned (true) or unpinned (false) saved takes."),
    sort: str = Query("newest", description="Sort mode: 'newest' (default), 'oldest', 'score', or 'pinned'."),
) -> dict:
    """List saved responses with optional search, filter, sort, pagination.

    Returns an envelope {items, total, page, per_page, total_pages, filters}
    so the UI can render pagination controls and a filter summary without
    inferring state. Free-tier users still see an empty list (not 403) —
    don't break the silent-gate contract that the existing /saved endpoint
    established.
    """
    # 60/min/user — list pagination scraping bound. Per-user so a single
    # abusive client can't starve other callers.
    enforce_user_rate_limit(
        user.id,
        scope="saved_list",
        limit=60,
        window_seconds=60,
        message="Too many saved-response list reads. Please slow down.",
    )
    if not has_feature(normalize_tier(get_tier_str(user)), "saved_responses"):
        return {
            "items": [],
            "total": 0,
            "page": 1,
            "per_page": per_page,
            "total_pages": 0,
            "filters": {"search": None, "persona_id": None, "min_score": None, "max_score": None, "pinned": None, "sort": "newest"},
        }

    q = db.query(SavedResponse).filter(SavedResponse.user_id == user.id)

    if search:
        # Bound via Query(max_length=100) AND sanitize again — defense in
        # depth so a malformed query string can't sneak a 10KB payload
        # through to the LIKE scan.
        safe = sanitize_model_optional_text(search, max_length=100, field_name="search")
        if safe:
            pattern = f"%{_escape_like(safe)}%"
            q = q.filter(
                or_(
                    SavedResponse.prompt.ilike(pattern, escape="\\"),
                    SavedResponse.one_liner.ilike(pattern, escape="\\"),
                )
            )

    if persona_id:
        # Exact match — persona_id is a closed enum string, not free text.
        q = q.filter(SavedResponse.persona_id == persona_id)

    if min_score is not None:
        q = q.filter(SavedResponse.score >= min_score)

    if max_score is not None:
        q = q.filter(SavedResponse.score <= max_score)

    if pinned is not None:
        if pinned:
            q = q.filter(SavedResponse.pinned_at.isnot(None))
        else:
            q = q.filter(SavedResponse.pinned_at.is_(None))

    # Sort. Unknown values fall back to newest so a stale frontend can't
    # break the endpoint; 'score' puts nulls last so untested takes don't
    # sink the top of "show me my best answers"
    if sort == "oldest":
        order_clauses = (SavedResponse.saved_at.asc(),)
    elif sort == "score":
        order_clauses = (
            SavedResponse.score.desc().nullslast(),
            SavedResponse.saved_at.desc(),
        )
    elif sort == "pinned":
        order_clauses = (
            SavedResponse.pinned_at.desc().nullslast(),
            SavedResponse.saved_at.desc(),
            SavedResponse.id.desc(),
        )
    else:
        order_clauses = (SavedResponse.saved_at.desc(),)

    total = q.count()
    rows = (
        q.order_by(*order_clauses)
        .offset((page - 1) * per_page)
        .limit(per_page)
        .all()
    )

    items = [
        {
            "id": row.id,
            "session_id": row.session_id,
            "agent_id": row.agent_id,
            "persona_id": row.persona_id,
            "persona_name": row.persona_name,
            "persona_color": row.persona_color,
            "prompt": row.prompt,
            "one_liner": row.one_liner,
            "verdict": row.verdict,
            "score": row.score,
            "confidence": row.confidence,
            "pinned": row.pinned_at is not None,
            "pinned_at": row.pinned_at.isoformat() if row.pinned_at else None,
            "saved_at": row.saved_at.isoformat() if row.saved_at else None,
        }
        for row in rows
    ]

    return {
        "items": items,
        "total": total,
        "page": page,
        "per_page": per_page,
        "total_pages": (total + per_page - 1) // per_page if per_page else 0,
        "filters": {
            "search": search,
            "persona_id": persona_id,
            "min_score": min_score,
            "max_score": max_score,
            "pinned": pinned,
            "sort": sort,
        },
    }


@router.patch("/saved/bulk-pin")
async def set_saved_bulk_pinned(
    body: BulkPinRequest,
    user: UserResponse = Depends(get_current_user_required),
    db: Session = Depends(get_db),
) -> dict:
    """Pin or unpin multiple saved takes at once.

    Ownership is scoped the same way as single pin/delete: rows that do not
    belong to the caller are silently ignored (no 403/404 existence oracle).
    Pinning respects SAVED_PIN_MAX; once the cap is reached, remaining
    unpinned rows are skipped and `pin_limit_reached` is set so the UI can
    tell the user why the bulk action was partial.
    """
    enforce_user_rate_limit(
        user.id,
        scope="saved_bulk_pin",
        limit=10,
        window_seconds=60,
        message="Too many saved-take bulk pin changes. Please slow down.",
    )
    if not has_feature(normalize_tier(get_tier_str(user)), "saved_responses"):
        raise HTTPException(
            status_code=403,
            detail={
                "error": "feature_not_allowed",
                "message": "Saved responses require a Plus or Pro subscription.",
                "upgrade_required": "plus",
            },
        )

    unique_ids = list(dict.fromkeys(body.ids))
    rows = (
        db.query(SavedResponse)
        .filter(
            SavedResponse.id.in_(unique_ids),
            SavedResponse.user_id == user.id,
        )
        .all()
    )
    by_id = {row.id: row for row in rows}

    if body.pinned:
        pinned_count = (
            db.query(SavedResponse.id)
            .filter(
                SavedResponse.user_id == user.id,
                SavedResponse.pinned_at.isnot(None),
            )
            .count()
        )
        unpinned_ids = [
            i for i in unique_ids if i in by_id and by_id[i].pinned_at is None
        ]
        remaining_slots = max(0, SAVED_PIN_MAX - int(pinned_count))
        to_pin = unpinned_ids[:remaining_slots]
        pin_limit_reached = len(to_pin) < len(unpinned_ids)
        for i in to_pin:
            by_id[i].pinned_at = utcnow_naive()
        already_pinned_ids = [
            i for i in unique_ids if i in by_id and by_id[i].pinned_at is not None
        ]
        applied_ids = to_pin + [
            i for i in already_pinned_ids if i not in set(to_pin)
        ]
    else:
        for i in unique_ids:
            if i in by_id:
                by_id[i].pinned_at = None
        applied_ids = [i for i in unique_ids if i in by_id]
        pin_limit_reached = False

    db.commit()
    return {
        "status": "ok",
        "requested": len(unique_ids),
        "applied": len(applied_ids),
        "ids": applied_ids,
        "pinned": body.pinned,
        "pin_limit_reached": pin_limit_reached,
    }


@router.patch("/saved/{saved_id}")
async def set_saved_pinned(
    saved_id: int,
    body: PinRequest,
    user: UserResponse = Depends(get_current_user_required),
    db: Session = Depends(get_db),
) -> dict:
    """Pin or unpin one saved take.

    Scope by owner so foreign IDs are indistinguishable from missing ones
    (same 404 contract as DELETE). Pinning is bounded to SAVED_PIN_MAX so
    a stale UI cannot turn every saved take into a pinned one silently.
    """
    enforce_user_rate_limit(
        user.id,
        scope="saved_pin",
        limit=60,
        window_seconds=3600,
        message="Too many saved-take pin changes. Limit is 60 per hour.",
    )
    if not has_feature(normalize_tier(get_tier_str(user)), "saved_responses"):
        raise HTTPException(
            status_code=403,
            detail={
                "error": "feature_not_allowed",
                "message": "Saved responses require a Plus or Pro subscription.",
                "upgrade_required": "plus",
            },
        )

    row = (
        db.query(SavedResponse)
        .filter(SavedResponse.id == saved_id, SavedResponse.user_id == user.id)
        .first()
    )
    if row is None:
        raise HTTPException(
            status_code=404,
            detail={"error": "not_found", "message": "Saved response not found"},
        )

    if body.pinned and row.pinned_at is None:
        pinned_count = (
            db.query(SavedResponse.id)
            .filter(
                SavedResponse.user_id == user.id,
                SavedResponse.pinned_at.isnot(None),
            )
            .count()
        )
        if pinned_count >= SAVED_PIN_MAX:
            raise HTTPException(
                status_code=400,
                detail={
                    "error": "pin_limit_reached",
                    "message": f"You can pin up to {SAVED_PIN_MAX} saved takes. Unpin one first.",
                },
            )
        row.pinned_at = utcnow_naive()
    elif not body.pinned:
        row.pinned_at = None

    db.commit()
    db.refresh(row)
    return {
        "id": row.id,
        "pinned": row.pinned_at is not None,
        "pinned_at": row.pinned_at.isoformat() if row.pinned_at else None,
    }


@router.post("/saved")
async def save_response(
    body: SavedRequest,
    user: UserResponse = Depends(get_current_user_required),
    db: Session = Depends(get_db),
) -> dict:
    if not has_feature(normalize_tier(get_tier_str(user)), "saved_responses"):
        raise HTTPException(
            status_code=403,
            detail={
                "error": "feature_not_allowed",
                "message": "Saved responses require a Plus or Pro subscription.",
                "upgrade_required": "plus",
            },
        )

    enforce_user_rate_limit(
        user.id,
        scope="saved_create",
        limit=60,
        window_seconds=3600,
        message="Too many saved takes. Limit is 60 per hour.",
    )

    existing = (
        db.query(SavedResponse)
        .filter(
            SavedResponse.user_id == user.id,
            SavedResponse.session_id == body.session_id,
            SavedResponse.agent_id == body.agent_id,
        )
        .first()
    )
    if existing:
        return {"status": "saved", "id": existing.id}

    count = (
        db.query(SavedResponse)
        .filter(SavedResponse.user_id == user.id)
        .count()
    )
    if int(count) >= SAVED_MAX_PER_USER:
        raise HTTPException(
            status_code=400,
            detail={
                "error": "saved_limit_reached",
                "message": (
                    f"Saved takes limit reached ({SAVED_MAX_PER_USER}). "
                    "Delete some before saving more."
                ),
                "active_cap": SAVED_MAX_PER_USER,
            },
        )

    row = SavedResponse(
        user_id=user.id,
        session_id=body.session_id,
        agent_id=body.agent_id,
        persona_id=body.persona_id,
        persona_name=body.persona_name,
        persona_color=body.persona_color,
        prompt=body.prompt,
        one_liner=body.one_liner,
        verdict=body.verdict,
        score=body.score,
        confidence=body.confidence,
    )
    db.add(row)
    db.commit()
    db.refresh(row)
    return {"status": "saved", "id": row.id}


@router.delete("/saved/bulk")
async def delete_saved_bulk(
    body: BulkDeleteRequest,
    user: UserResponse = Depends(get_current_user_required),
    db: Session = Depends(get_db),
) -> dict:
    """Bulk delete — for the 'select all visible' cleanup pattern.

    Foreign ids (not owned by the caller) are silently dropped from the
    delete set. The response reports requested / deleted counts so the UI
    can show a partial-success message if a stale page referenced ids
    another user has since claimed.
    """
    # 10/min/user — destructive bulk delete; ownership is gated but a hostile
    # caller can still burn DB hits if uncapped. Same shape as DELETE rooms (40).
    enforce_user_rate_limit(
        user.id,
        scope="saved_bulk_delete",
        limit=10,
        window_seconds=60,
        message="Too many saved-response bulk delete attempts. Please slow down.",
    )
    if not has_feature(normalize_tier(get_tier_str(user)), "saved_responses"):
        raise HTTPException(
            status_code=403,
            detail={
                "error": "feature_not_allowed",
                "message": "Saved responses require a Plus or Pro subscription.",
                "upgrade_required": "plus",
            },
        )

    # Deduplicate within the request — a UI bug that double-fires the same
    # id shouldn't double-count in the response.
    unique_ids = list(dict.fromkeys(body.ids))
    requested = len(unique_ids)

    # Scope by owner so we never delete another user's rows even if a UI
    # bug hands us foreign ids. Return the exact owned ids so the client can
    # reconcile its local library when a mixed/partial request succeeds.
    owned_ids = [
        row_id
        for (row_id,) in db.query(SavedResponse.id)
        .filter(
            SavedResponse.id.in_(unique_ids),
            SavedResponse.user_id == user.id,
        )
        .all()
    ]
    deleted = 0
    if owned_ids:
        deleted = (
            db.query(SavedResponse)
            .filter(SavedResponse.id.in_(owned_ids))
            .delete(synchronize_session=False)
        )
        db.commit()
    return {
        "status": "deleted",
        "requested": requested,
        "deleted": int(deleted or 0),
        "ids": owned_ids,
    }


@router.delete("/saved/{saved_id}")
async def delete_saved(
    saved_id: int,
    user: UserResponse = Depends(get_current_user_required),
    db: Session = Depends(get_db),
) -> dict:
    # 10/min/user — destructive single delete. Same shape as DELETE rooms /bulk.
    enforce_user_rate_limit(
        user.id,
        scope="saved_delete",
        limit=10,
        window_seconds=60,
        message="Too many saved-response delete attempts. Please slow down.",
    )
    if not has_feature(normalize_tier(get_tier_str(user)), "saved_responses"):
        raise HTTPException(
            status_code=403,
            detail={
                "error": "feature_not_allowed",
                "message": "Saved responses require a Plus or Pro subscription.",
                "upgrade_required": "plus",
            },
        )

    # Scope by owner so foreign IDs are indistinguishable from missing ones
    # (no 403 existence oracle).
    row = (
        db.query(SavedResponse)
        .filter(SavedResponse.id == saved_id, SavedResponse.user_id == user.id)
        .first()
    )
    if row is None:
        raise HTTPException(
            status_code=404,
            detail={"error": "not_found", "message": "Saved response not found"},
        )
    db.delete(row)
    db.commit()
    return {"status": "deleted", "id": saved_id}


@router.get("/saved/export")
async def export_saved(
    user: UserResponse = Depends(get_current_user_required),
    db: Session = Depends(get_db),
    search: Optional[str] = Query(None, max_length=100, description="Case-insensitive substring match on prompt + one_liner."),
    persona_id: Optional[str] = Query(None, max_length=50, description="Restrict to one persona."),
    min_score: Optional[int] = Query(None, ge=0, le=100, description="Minimum score (inclusive)."),
    max_score: Optional[int] = Query(None, ge=0, le=100, description="Maximum score (inclusive)."),
    pinned: Optional[bool] = Query(None, description="Restrict to pinned (true) or unpinned (false) saved takes."),
    sort: str = Query("newest", description="Sort mode: 'newest' (default), 'oldest', or 'score'."),
    format: str = Query("csv", description="Export format: 'csv' (default), 'json', or 'xlsx'."),
):
    """Export saved responses in CSV, JSON, or XLSX format.

    Supports the same filters as /api/saved plus min_score, max_score, and
    pinned (true = pinned only, false = unpinned only).
    CSV format includes formula-injection defense.
    JSON format provides structured data for programmatic use.
    XLSX format provides Excel-compatible spreadsheets.
    """
    enforce_user_rate_limit(
        user.id,
        scope="saved_export",
        limit=30,
        window_seconds=60,
        message="Too many exports. Please wait.",
    )

    if not has_feature(normalize_tier(get_tier_str(user)), "saved_responses"):
        raise HTTPException(
            status_code=403,
            detail={"error": "feature_not_allowed", "message": "Saved responses require Plus or Pro."},
        )

    # Normalize once so disclosed filters always match the query actually run
    # (same parity contract as the preset preview endpoint).
    safe_search = normalize_export_search(search)

    # Build query with same filters as get_saved (shared with preset preview)
    q = build_saved_export_query(
        db,
        user.id,
        search=safe_search,
        persona_id=persona_id,
        min_score=min_score,
        max_score=max_score,
        sort=sort,
        pinned=pinned,
    )

    saved_items = q.all()

    def _csv_safe(value) -> str:
        """Escape value for CSV to prevent formula injection."""
        if value is None:
            return ""
        s = str(value)
        if s.startswith(("=", "+", "-", "@")):
            return "'" + s
        return s

    from arena.core.datetime_utils import utcnow_naive
    from fastapi.responses import Response
    from arena.core.http_headers import content_disposition_attachment

    export_timestamp = utcnow_naive()

    if format == "json":
        # JSON export format
        items = []
        for item in saved_items:
            items.append({
                "id": item.id,
                "session_id": item.session_id,
                "agent_id": item.agent_id,
                "persona_id": item.persona_id,
                "persona_name": item.persona_name,
                "persona_color": item.persona_color,
                "prompt": item.prompt,
                "one_liner": item.one_liner,
                "verdict": item.verdict,
                "score": item.score,
                "confidence": item.confidence,
                "saved_at": item.saved_at.isoformat() if item.saved_at else None,
            })

        export_data = {
            "metadata": {
                "export_format": "json",
                "exported_at": export_timestamp.isoformat(),
                "total_count": len(items),
                "filters": {
                    "search": safe_search,
                    "persona_id": persona_id,
                    "min_score": min_score,
                    "max_score": max_score,
                    "pinned": pinned,
                    "sort": sort,
                },
            },
            "data": items,
        }

        filename = f"arena-saved-{user.id}-{export_timestamp.strftime('%Y%m%d-%H%M%S')}.json"
        headers = {
            "Content-Disposition": content_disposition_attachment(filename),
            "X-Content-Type-Options": "nosniff",
            "Cache-Control": "no-store, no-cache, must-revalidate, private",
        }
        return Response(
            content=json.dumps(export_data, indent=2, default=str),
            media_type="application/json; charset=utf-8",
            headers=headers,
        )

    elif format == "xlsx":
        # XLSX export format
        if not OPENPYXL_AVAILABLE:
            raise HTTPException(
                status_code=503,
                detail={
                    "error": "xlsx_export_unavailable",
                    "message": "XLSX export requires openpyxl package. Please install it.",
                },
            )

        wb = Workbook()

        # Add Summary sheet first
        summary_ws = wb.active
        summary_ws.title = "Summary"

        # Summary information
        summary_ws.append(["Arena Saved Responses Export"])
        summary_ws.append([""])
        summary_ws.append(["Export Details:"])
        summary_ws.append(["Format:", "XLSX"])
        summary_ws.append(["Exported At:", export_timestamp.strftime('%Y-%m-%d %H:%M:%S UTC')])
        summary_ws.append(["Total Records:", len(saved_items)])
        summary_ws.append([""])
        summary_ws.append(["Filters Applied:"])
        summary_ws.append(["Search:", safe_search or "None"])
        summary_ws.append(["Persona:", persona_id or "All"])
        summary_ws.append(["Min Score:", str(min_score) if min_score is not None else "None"])
        summary_ws.append(["Max Score:", str(max_score) if max_score is not None else "None"])
        summary_ws.append(["Pinned:", "pinned only" if pinned is True else ("unpinned only" if pinned is False else "All")])
        summary_ws.append(["Sort:", sort])
        summary_ws.append([""])
        summary_ws.append(["User ID:", user.id])

        # Style summary sheet
        from openpyxl.styles import Font, Alignment, Border, Side, PatternFill
        from openpyxl.styles.colors import Color

        bold_font = Font(bold=True)
        gray_fill = PatternFill(start_color="F2F2F2", end_color="F2F2F2", fill_type="solid")
        thin_border = Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'), bottom=Side(style='thin'))

        # Style summary header
        summary_ws["A1"].font = Font(bold=True, size=14, color=Color("0066CC"))
        summary_ws["A1"].alignment = Alignment(horizontal='center')

        # Style key-value pairs
        for row in summary_ws.iter_rows(min_row=3, max_row=14, min_col=1, max_col=2):
            for cell in row:
                cell.border = thin_border
                if cell.column == 1:  # Key column
                    cell.font = bold_font
                    cell.fill = gray_fill

        # Set column widths for summary
        summary_ws.column_dimensions["A"].width = 20
        summary_ws.column_dimensions["B"].width = 30

        # Add Data sheet
        data_ws = wb.create_sheet(title="Data")

        # Write header
        headers_row = [
            "ID", "Session ID", "Agent ID", "Persona ID", "Persona Name", "Persona Color",
            "Prompt", "One Liner", "Verdict", "Score", "Confidence", "Saved At"
        ]
        data_ws.append(headers_row)

        # Style header row
        header_fill = PatternFill(start_color="4F81BD", end_color="4F81BD", fill_type="solid")
        header_font = Font(bold=True, color=Color("FFFFFF"))
        header_alignment = Alignment(horizontal='center')

        for cell in data_ws[1]:
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = header_alignment
            cell.border = thin_border

        # Write data rows
        for item in saved_items:
            row = [
                item.id,
                item.session_id,
                item.agent_id,
                item.persona_id,
                item.persona_name,
                item.persona_color,
                item.prompt[:500] if item.prompt else "",  # Truncate long prompts
                item.one_liner,
                item.verdict[:500] if item.verdict else "",  # Truncate long verdicts
                item.score,
                item.confidence,
                item.saved_at.isoformat() if item.saved_at else "",
            ]
            data_ws.append(row)

        # Style data cells
        for row in data_ws.iter_rows(min_row=2):  # Skip header
            for cell in row:
                cell.border = thin_border
                # Right align numbers
                if cell.column in [9, 10, 11]:  # Score, Confidence, Saved At
                    cell.alignment = Alignment(horizontal='right')
                # Left align text
                else:
                    cell.alignment = Alignment(horizontal='left', wrap_text=True)

        # Auto-adjust column widths for data sheet
        for col in data_ws.columns:
            max_length = 0
            column = col[0].column_letter  # Get the column name
            for cell in col:
                try:
                    cell_length = len(str(cell.value)) if cell.value else 0
                    if cell_length > max_length:
                        max_length = cell_length
                except:
                    logger.debug("Cell length measurement failed", exc_info=True)
            adjusted_width = (max_length + 2) * 1.2
            data_ws.column_dimensions[column].width = max(10, min(adjusted_width, 80))  # Cap at 80

        # Freeze header row
        data_ws.freeze_panes = "A2"

        # Set data sheet as active (more intuitive for users)
        wb.active = data_ws

        # Save workbook to bytes
        xlsx_buffer = io.BytesIO()
        wb.save(xlsx_buffer)
        xlsx_buffer.seek(0)

        filename = f"arena-saved-{user.id}-{export_timestamp.strftime('%Y%m%d-%H%M%S')}.xlsx"
        headers = {
            "Content-Disposition": content_disposition_attachment(filename),
            "X-Content-Type-Options": "nosniff",
            "Cache-Control": "no-store, no-cache, must-revalidate, private",
        }
        return Response(
            content=xlsx_buffer.getvalue(),
            media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            headers=headers,
        )

    # CSV export format (default)
    buf = io.StringIO()
    writer = csv.writer(buf, quoting=csv.QUOTE_MINIMAL, lineterminator="\r\n")

    # Write header
    writer.writerow([
        "id",
        "session_id",
        "agent_id",
        "persona_id",
        "persona_name",
        "persona_color",
        "prompt",
        "one_liner",
        "verdict",
        "score",
        "confidence",
        "saved_at",
    ])

    # Write rows
    for item in saved_items:
        writer.writerow([
            _csv_safe(item.id),
            _csv_safe(item.session_id),
            _csv_safe(item.agent_id),
            _csv_safe(item.persona_id),
            _csv_safe(item.persona_name),
            _csv_safe(item.persona_color),
            _csv_safe(item.prompt[:500] if item.prompt else ""),  # Truncate long prompts
            _csv_safe(item.one_liner),
            _csv_safe(item.verdict[:500] if item.verdict else ""),  # Truncate long verdicts
            _csv_safe(item.score),
            _csv_safe(item.confidence),
            _csv_safe(item.saved_at.isoformat() if item.saved_at else ""),
        ])

    filename = f"arena-saved-{user.id}-{export_timestamp.strftime('%Y%m%d-%H%M%S')}.csv"
    headers = {
        "Content-Disposition": content_disposition_attachment(filename),
        "X-Content-Type-Options": "nosniff",
        "Cache-Control": "no-store, no-cache, must-revalidate, private",
    }
    return Response(
        content=buf.getvalue(),
        media_type="text/csv; charset=utf-8",
        headers=headers,
    )


# Keep backward compatibility with old CSV-only endpoint
@router.get("/saved/export.csv")
async def export_saved_csv_legacy(
    user: UserResponse = Depends(get_current_user_required),
    db: Session = Depends(get_db),
    search: Optional[str] = Query(None, max_length=100),
    persona_id: Optional[str] = Query(None, max_length=50),
    min_score: Optional[int] = Query(None, ge=0, le=100),
    pinned: Optional[bool] = Query(None),
    sort: str = Query("newest"),
):
    """Legacy CSV-only export endpoint - redirects to new unified export with format=csv."""
    from fastapi.responses import RedirectResponse

    # Mirror the unified export throttle so the redirect can't be used to
    # bypass the per-user export rate limit.
    enforce_user_rate_limit(
        user.id,
        scope="saved_export",
        limit=30,
        window_seconds=60,
        message="Too many exports. Please wait.",
    )

    # Build query parameters for redirect
    params = []
    if search:
        params.append(f"search={search}")
    if persona_id:
        params.append(f"persona_id={persona_id}")
    if min_score is not None:
        params.append(f"min_score={min_score}")
    if pinned is not None:
        params.append(f"pinned={'true' if pinned else 'false'}")
    if sort != "newest":
        params.append(f"sort={sort}")
    params.append("format=csv")

    query_string = "&".join(params) if params else "format=csv"
    return RedirectResponse(url=f"/api/saved/export?{query_string}", status_code=307)
