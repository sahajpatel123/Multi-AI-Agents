import io
import pytest
from arena.core.datetime_utils import utcnow_naive
from arena.core.auth import create_access_token
from arena.db_models import SavedResponse, UserTier


def _make_pro(make_user):
    return make_user(email="pro_saved@example.com", tier=UserTier.PRO)


def _pro_headers(user):
    """Build the Authorization header for a pro-tier user."""
    return {"Authorization": f"Bearer {create_access_token(user.id, user.email)}"}


def _seed_saved(
    db_session,
    user_id,
    saved_id,
    prompt="Test prompt",
    one_liner="Test answer",
    score=85,
    persona_id=None,
    pinned_at=None,
):
    saved = SavedResponse(
        user_id=user_id,
        session_id=f"sess-{saved_id}",
        agent_id=f"agent-{saved_id}",
        persona_id=persona_id or f"persona-{saved_id}",
        persona_name=f"Persona {saved_id}",
        persona_color="blue",
        prompt=prompt,
        one_liner=one_liner,
        verdict=f"This is the verdict for {saved_id}",
        score=score,
        confidence=90,
        saved_at=utcnow_naive(),
        pinned_at=pinned_at,
    )
    db_session.add(saved)
    db_session.flush()
    return saved


@pytest.mark.asyncio
async def test_saved_csv_export(app_client, make_user, db_session):
    """Test CSV export of saved responses using new unified endpoint."""
    user = _make_pro(make_user)
    db_session.commit()

    _seed_saved(db_session, user.id, "save-1", prompt="Bitcoin question", score=90)
    _seed_saved(db_session, user.id, "save-2", prompt="Ethereum question", score=85)
    db_session.commit()

    res = await app_client.get(
        "/api/saved/export?format=csv",
        headers=_pro_headers(user),
    )
    assert res.status_code == 200
    assert "text/csv" in res.headers["content-type"]
    text = res.text
    assert "id" in text
    assert "session_id" in text
    assert "prompt" in text
    assert "one_liner" in text
    assert "verdict" in text
    assert "score" in text
    assert "persona_color" in text  # Added in polish
    assert "sess-save-1" in text
    assert "sess-save-2" in text


@pytest.mark.asyncio
async def test_saved_csv_with_search_filter(app_client, make_user, db_session):
    """Test CSV export with search filter."""
    user = _make_pro(make_user)
    db_session.commit()

    _seed_saved(db_session, user.id, "save-1", prompt="Bitcoin analysis", one_liner="Bitcoin is up")
    _seed_saved(db_session, user.id, "save-2", prompt="Ethereum analysis", one_liner="Ethereum is down")
    _seed_saved(db_session, user.id, "save-3", prompt="Bitcoin forecast", one_liner="Bitcoin will rise")
    db_session.commit()

    res = await app_client.get(
        "/api/saved/export?format=csv&search=Bitcoin",
        headers=_pro_headers(user),
    )
    assert res.status_code == 200
    text = res.text
    assert "sess-save-1" in text
    assert "sess-save-3" in text
    assert "sess-save-2" not in text


@pytest.mark.asyncio
async def test_saved_csv_with_persona_filter(app_client, make_user, db_session):
    """Test CSV export with persona filter."""
    user = _make_pro(make_user)
    db_session.commit()

    _seed_saved(db_session, user.id, "save-1", persona_id="analyst")
    _seed_saved(db_session, user.id, "save-2", persona_id="researcher")
    _seed_saved(db_session, user.id, "save-3", persona_id="analyst")
    db_session.commit()

    res = await app_client.get(
        "/api/saved/export?format=csv&persona_id=analyst",
        headers=_pro_headers(user),
    )
    assert res.status_code == 200
    text = res.text
    assert "sess-save-1" in text
    assert "sess-save-3" in text
    assert "sess-save-2" not in text


@pytest.mark.asyncio
async def test_saved_csv_with_min_score_filter(app_client, make_user, db_session):
    """Test CSV export with minimum score filter."""
    user = _make_pro(make_user)
    db_session.commit()

    _seed_saved(db_session, user.id, "save-1", score=90)
    _seed_saved(db_session, user.id, "save-2", score=80)
    _seed_saved(db_session, user.id, "save-3", score=85)
    db_session.commit()

    res = await app_client.get(
        "/api/saved/export?format=csv&min_score=85",
        headers=_pro_headers(user),
    )
    assert res.status_code == 200
    text = res.text
    assert "sess-save-1" in text
    assert "sess-save-3" in text
    assert "sess-save-2" not in text


@pytest.mark.asyncio
async def test_saved_csv_with_sort(app_client, make_user, db_session):
    """Test CSV export with sort parameter."""
    user = _make_pro(make_user)
    db_session.commit()

    _seed_saved(db_session, user.id, "save-1", score=90)
    _seed_saved(db_session, user.id, "save-2", score=80)
    _seed_saved(db_session, user.id, "save-3", score=85)
    db_session.commit()

    # Test score sort (descending)
    res = await app_client.get(
        "/api/saved/export?format=csv&sort=score",
        headers=_pro_headers(user),
    )
    assert res.status_code == 200
    # The order in CSV should have highest score first
    text = res.text
    assert "id" in text
    assert "sess-save-1" in text


@pytest.mark.asyncio
async def test_saved_csv_formula_injection_defense(app_client, make_user, db_session):
    """Test CSV export defends against formula injection."""
    user = _make_pro(make_user)
    db_session.commit()

    _seed_saved(db_session, user.id, "formula-test", prompt="=SUM(A1:B1)")
    db_session.commit()

    res = await app_client.get(
        "/api/saved/export?format=csv",
        headers=_pro_headers(user),
    )
    assert res.status_code == 200
    text = res.text
    # Formula should be quoted/escaped
    assert "'=SUM(A1:B1)" in text or "=SUM" not in text


@pytest.mark.asyncio
async def test_saved_csv_empty(app_client, make_user, db_session):
    """Test CSV export when user has no saved responses."""
    user = _make_pro(make_user)
    db_session.commit()

    res = await app_client.get(
        "/api/saved/export?format=csv",
        headers=_pro_headers(user),
    )
    assert res.status_code == 200
    text = res.text
    assert "id" in text  # Header should be present
    assert "persona_color" in text  # New field from polish
    assert "sess-save" not in text  # No data rows


@pytest.mark.asyncio
async def test_saved_csv_403_for_guest(app_client, make_user, db_session):
    """Test that guest users without access get 403."""
    from arena.db_models import UserTier as DBUserTier
    user = make_user(email="guest_saved@example.com", tier=DBUserTier.GUEST)
    db_session.commit()

    res = await app_client.get(
        "/api/saved/export?format=csv",
        headers={"Authorization": f"Bearer {create_access_token(user.id, user.email)}"},
    )
    # Guest users should get 403 (Forbidden) - saved responses require Plus/Pro
    assert res.status_code == 403


# JSON Export Tests (added in Loop 15 - POLISH phase)
@pytest.mark.asyncio
async def test_saved_json_export(app_client, make_user, db_session):
    """Test JSON export of saved responses."""
    user = _make_pro(make_user)
    db_session.commit()

    _seed_saved(db_session, user.id, "save-1", prompt="Bitcoin question", score=90)
    _seed_saved(db_session, user.id, "save-2", prompt="Ethereum question", score=85)
    db_session.commit()

    res = await app_client.get(
        "/api/saved/export?format=json",
        headers=_pro_headers(user),
    )
    assert res.status_code == 200
    assert "application/json" in res.headers["content-type"]

    data = res.json()
    assert "metadata" in data
    assert "data" in data
    assert data["metadata"]["export_format"] == "json"
    assert data["metadata"]["total_count"] == 2
    assert "exported_at" in data["metadata"]
    assert len(data["data"]) == 2

    # Check first item structure
    item = data["data"][0]
    assert "id" in item
    assert "session_id" in item
    assert "prompt" in item
    assert "persona_color" in item  # Should be included in JSON


@pytest.mark.asyncio
async def test_saved_json_export_with_filters(app_client, make_user, db_session):
    """Test JSON export with filters."""
    user = _make_pro(make_user)
    db_session.commit()

    _seed_saved(db_session, user.id, "save-1", prompt="Bitcoin analysis", persona_id="analyst")
    _seed_saved(db_session, user.id, "save-2", prompt="Ethereum analysis", persona_id="researcher")
    db_session.commit()

    res = await app_client.get(
        "/api/saved/export?format=json&persona_id=analyst",
        headers=_pro_headers(user),
    )
    assert res.status_code == 200

    data = res.json()
    assert data["metadata"]["total_count"] == 1
    assert data["metadata"]["filters"]["persona_id"] == "analyst"
    assert len(data["data"]) == 1
    assert data["data"][0]["persona_id"] == "analyst"


@pytest.mark.asyncio
async def test_saved_csv_pinned_filter(app_client, make_user, db_session):
    """CSV export with pinned=true must only include pinned takes."""
    user = _make_pro(make_user)
    db_session.commit()

    _seed_saved(
        db_session,
        user.id,
        "save-1",
        prompt="pinned row",
        pinned_at=utcnow_naive(),
    )
    _seed_saved(db_session, user.id, "save-2", prompt="unpinned row")
    db_session.commit()

    res = await app_client.get(
        "/api/saved/export?format=csv&pinned=true",
        headers=_pro_headers(user),
    )
    assert res.status_code == 200
    text = res.text
    assert "pinned row" in text
    assert "unpinned row" not in text


@pytest.mark.asyncio
async def test_saved_json_export_discloses_pinned_filter(
    app_client, make_user, db_session
):
    """JSON export metadata should disclose the pinned filter applied."""
    user = _make_pro(make_user)
    db_session.commit()

    _seed_saved(
        db_session,
        user.id,
        "save-1",
        prompt="pinned row",
        pinned_at=utcnow_naive(),
    )
    _seed_saved(db_session, user.id, "save-2", prompt="unpinned row")
    db_session.commit()

    res = await app_client.get(
        "/api/saved/export?format=json&pinned=true",
        headers=_pro_headers(user),
    )
    assert res.status_code == 200
    body = res.json()
    assert body["metadata"]["filters"]["pinned"] is True
    assert {row["prompt"] for row in body["data"]} == {"pinned row"}


@pytest.mark.asyncio
async def test_saved_json_export_whitespace_search_degrades_to_none(
    app_client, make_user, db_session
):
    """Whitespace-only search must not 500 and must disclose as None (parity
    with the preset preview contract: disclosed filters == query actually run).
    """
    user = _make_pro(make_user)
    db_session.commit()

    _seed_saved(db_session, user.id, "save-1", prompt="Bitcoin question")
    _seed_saved(db_session, user.id, "save-2", prompt="Ethereum question")
    db_session.commit()

    res = await app_client.get(
        "/api/saved/export?format=json&search=%20%20%20",
        headers=_pro_headers(user),
    )
    assert res.status_code == 200
    data = res.json()
    assert data["metadata"]["total_count"] == 2
    assert data["metadata"]["filters"]["search"] is None
    assert len(data["data"]) == 2


@pytest.mark.asyncio
async def test_saved_json_export_empty(app_client, make_user, db_session):
    """Test JSON export when user has no saved responses."""
    user = _make_pro(make_user)
    db_session.commit()

    res = await app_client.get(
        "/api/saved/export?format=json",
        headers=_pro_headers(user),
    )
    assert res.status_code == 200
    data = res.json()
    assert data["metadata"]["total_count"] == 0
    assert data["data"] == []


@pytest.mark.asyncio
async def test_saved_json_export_403_for_guest(app_client, make_user, db_session):
    """Test that guest users get 403 for JSON export."""
    from arena.db_models import UserTier as DBUserTier
    user = make_user(email="guest_json@example.com", tier=DBUserTier.GUEST)
    db_session.commit()

    res = await app_client.get(
        "/api/saved/export?format=json",
        headers={"Authorization": f"Bearer {create_access_token(user.id, user.email)}"},
    )
    assert res.status_code == 403


@pytest.mark.asyncio
async def test_saved_export_default_format_is_csv(app_client, make_user, db_session):
    """Test that default format is CSV when not specified."""
    user = _make_pro(make_user)
    db_session.commit()

    _seed_saved(db_session, user.id, "save-1", prompt="Test")
    db_session.commit()

    res = await app_client.get(
        "/api/saved/export",
        headers=_pro_headers(user),
    )
    assert res.status_code == 200
    assert "text/csv" in res.headers["content-type"]


@pytest.mark.asyncio
async def test_saved_export_filename_has_timestamp(app_client, make_user, db_session):
    """Test that export filename includes timestamp for uniqueness."""
    user = _make_pro(make_user)
    db_session.commit()

    _seed_saved(db_session, user.id, "save-1", prompt="Test")
    db_session.commit()

    res = await app_client.get(
        "/api/saved/export?format=csv",
        headers=_pro_headers(user),
    )
    assert res.status_code == 200
    content_disposition = res.headers["content-disposition"]
    # Should contain timestamp pattern like 20260804-145600
    assert "-" in content_disposition
    assert ".csv" in content_disposition

    res_json = await app_client.get(
        "/api/saved/export?format=json",
        headers=_pro_headers(user),
    )
    assert res_json.status_code == 200
    content_disposition_json = res_json.headers["content-disposition"]
    assert "-" in content_disposition_json
    assert ".json" in content_disposition_json


# XLSX Export Tests (added in Loop 16 - ADD phase)
@pytest.mark.asyncio
async def test_saved_xlsx_export(app_client, make_user, db_session):
    """Test XLSX export of saved responses."""
    user = _make_pro(make_user)
    db_session.commit()

    _seed_saved(db_session, user.id, "save-1", prompt="Bitcoin question", score=90)
    _seed_saved(db_session, user.id, "save-2", prompt="Ethereum question", score=85)
    db_session.commit()

    res = await app_client.get(
        "/api/saved/export?format=xlsx",
        headers=_pro_headers(user),
    )
    assert res.status_code == 200
    assert "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet" in res.headers["content-type"]

    # Verify it's a valid XLSX file by checking magic bytes
    content = res.content
    assert len(content) > 0
    # XLSX files should start with PK magic number (ZIP format)
    assert content[:2] == b'PK'

    # Verify filename
    content_disposition = res.headers["content-disposition"]
    assert ".xlsx" in content_disposition


@pytest.mark.asyncio
async def test_saved_xlsx_export_with_filters(app_client, make_user, db_session):
    """Test XLSX export with filters."""
    user = _make_pro(make_user)
    db_session.commit()

    _seed_saved(db_session, user.id, "save-1", prompt="Bitcoin analysis", persona_id="analyst")
    _seed_saved(db_session, user.id, "save-2", prompt="Ethereum analysis", persona_id="researcher")
    db_session.commit()

    res = await app_client.get(
        "/api/saved/export?format=xlsx&persona_id=analyst",
        headers=_pro_headers(user),
    )
    assert res.status_code == 200
    assert ".xlsx" in res.headers["content-disposition"]

    # Verify it's valid XLSX
    content = res.content
    assert content[:2] == b'PK'


@pytest.mark.asyncio
async def test_saved_xlsx_export_empty(app_client, make_user, db_session):
    """Test XLSX export when user has no saved responses."""
    user = _make_pro(make_user)
    db_session.commit()

    res = await app_client.get(
        "/api/saved/export?format=xlsx",
        headers=_pro_headers(user),
    )
    assert res.status_code == 200
    content = res.content
    assert content[:2] == b'PK'  # Still valid XLSX with just headers


@pytest.mark.asyncio
async def test_saved_xlsx_export_403_for_guest(app_client, make_user, db_session):
    """Test that guest users get 403 for XLSX export."""
    from arena.db_models import UserTier as DBUserTier
    user = make_user(email="guest_xlsx@example.com", tier=DBUserTier.GUEST)
    db_session.commit()

    res = await app_client.get(
        "/api/saved/export?format=xlsx",
        headers={"Authorization": f"Bearer {create_access_token(user.id, user.email)}"},
    )
    assert res.status_code == 403


@pytest.mark.asyncio
async def test_saved_xlsx_export_filename_has_timestamp(app_client, make_user, db_session):
    """Test that XLSX export filename includes timestamp."""
    user = _make_pro(make_user)
    db_session.commit()

    _seed_saved(db_session, user.id, "save-1", prompt="Test")
    db_session.commit()

    res = await app_client.get(
        "/api/saved/export?format=xlsx",
        headers=_pro_headers(user),
    )
    assert res.status_code == 200
    content_disposition = res.headers["content-disposition"]
    assert "-" in content_disposition
    assert ".xlsx" in content_disposition


# Enhanced XLSX Tests (added in Loop 17 - POLISH phase)
@pytest.mark.asyncio
async def test_saved_xlsx_has_multiple_sheets(app_client, make_user, db_session):
    """Test that XLSX export has Summary and Data sheets."""
    try:
        import openpyxl
    except ImportError:
        pytest.skip("openpyxl not available")

    user = _make_pro(make_user)
    db_session.commit()

    _seed_saved(db_session, user.id, "save-1", prompt="Test prompt")
    db_session.commit()

    res = await app_client.get(
        "/api/saved/export?format=xlsx",
        headers=_pro_headers(user),
    )
    assert res.status_code == 200

    # Load the XLSX file and check sheets
    wb = openpyxl.load_workbook(io.BytesIO(res.content))
    sheet_names = wb.sheetnames

    assert "Summary" in sheet_names
    assert "Data" in sheet_names

    # Check summary sheet content
    summary_ws = wb["Summary"]
    summary_content = [[cell.value for cell in row] for row in summary_ws.iter_rows()]

    # Should contain export details
    assert any("Arena Saved Responses Export" in str(row) for row in summary_content)
    assert any("XLSX" in str(row) for row in summary_content)
    assert any("Total Records:" in str(row) for row in summary_content)


@pytest.mark.asyncio
async def test_saved_xlsx_summary_includes_max_score_filter(
    app_client, make_user, db_session
):
    """XLSX summary discloses the same filters as JSON/preview, incl. max_score."""
    try:
        import openpyxl
    except ImportError:
        pytest.skip("openpyxl not available")

    user = _make_pro(make_user)
    db_session.commit()

    _seed_saved(db_session, user.id, "save-1", prompt="Test prompt", score=90)
    db_session.commit()

    res = await app_client.get(
        "/api/saved/export?format=xlsx&min_score=80&max_score=95",
        headers=_pro_headers(user),
    )
    assert res.status_code == 200

    wb = openpyxl.load_workbook(io.BytesIO(res.content))
    summary_ws = wb["Summary"]
    summary_content = [[cell.value for cell in row] for row in summary_ws.iter_rows()]
    flat = [str(cell) for row in summary_content for cell in row if cell is not None]

    assert "Max Score:" in flat
    assert "95" in flat
    assert "Min Score:" in flat
    assert "80" in flat


@pytest.mark.asyncio
async def test_saved_xlsx_has_styled_headers(app_client, make_user, db_session):
    """Test that XLSX data sheet has styled headers."""
    try:
        import openpyxl
    except ImportError:
        pytest.skip("openpyxl not available")

    user = _make_pro(make_user)
    db_session.commit()

    _seed_saved(db_session, user.id, "save-1", prompt="Test prompt")
    db_session.commit()

    res = await app_client.get(
        "/api/saved/export?format=xlsx",
        headers=_pro_headers(user),
    )
    assert res.status_code == 200

    # Load and check header styling
    wb = openpyxl.load_workbook(io.BytesIO(res.content))
    data_ws = wb["Data"]

    # Header should be in row 1
    header_cell = data_ws["A1"]
    assert header_cell.value == "ID"

    # Check that header has bold font and fill (check attributes directly)
    assert hasattr(header_cell.font, 'bold')
    assert header_cell.font.bold
    assert hasattr(header_cell.fill, 'start_color')
    assert header_cell.fill.start_color.rgb is not None


@pytest.mark.asyncio
async def test_saved_xlsx_has_frozen_panes(app_client, make_user, db_session):
    """Test that XLSX data sheet has frozen header row."""
    try:
        import openpyxl
    except ImportError:
        pytest.skip("openpyxl not available")

    user = _make_pro(make_user)
    db_session.commit()

    _seed_saved(db_session, user.id, "save-1", prompt="Test prompt")
    db_session.commit()

    res = await app_client.get(
        "/api/saved/export?format=xlsx",
        headers=_pro_headers(user),
    )
    assert res.status_code == 200

    # Load and check frozen panes
    wb = openpyxl.load_workbook(io.BytesIO(res.content))
    data_ws = wb["Data"]

    # Should have frozen panes at A2 (header row frozen)
    assert data_ws.freeze_panes == "A2"
